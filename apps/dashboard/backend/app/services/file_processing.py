import csv
import io
import logging
from dataclasses import dataclass
from itertools import chain
from pathlib import Path
from typing import Iterable, List, Optional, Sequence

import openpyxl
import xlrd
import xlwt

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
OUTPUT_COLUMNS = ["verification_status", "is_role_based", "validated_at"]


class FileProcessingError(Exception):
    def __init__(self, message: str, *, details: Optional[dict] = None):
        super().__init__(message)
        self.details = details or {}


@dataclass
class ParsedEmails:
    emails: List[str]
    email_column_index: int


def _column_letters_to_index(value: str) -> Optional[int]:
    trimmed = value.strip()
    if not trimmed:
        return None
    if trimmed.isdigit():
        index = int(trimmed) - 1
        return index if index >= 0 else None
    if not trimmed.isalpha():
        return None
    total = 0
    for ch in trimmed.upper():
        total = total * 26 + (ord(ch) - ord("A") + 1)
    return total - 1 if total > 0 else None


def _resolve_column_index(headers: Sequence[str], email_column: str, first_row_has_labels: bool) -> int:
    index = _column_letters_to_index(email_column)
    if index is not None:
        return index
    if not first_row_has_labels:
        raise FileProcessingError("Email column must be a column letter (e.g., A) or number when no headers exist")
    target = email_column.strip().lower()
    if not target:
        raise FileProcessingError("Email column name is required")
    matches = [i for i, header in enumerate(headers) if header.strip().lower() == target]
    if not matches:
        raise FileProcessingError(
            "Email column header not found",
            details={"available_headers": [header for header in headers if header.strip()]},
        )
    if len(matches) > 1:
        raise FileProcessingError("Email column header is not unique", details={"matches": matches})
    return matches[0]


def _normalize_email(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _collect_emails(
    rows: Iterable[Sequence[object]],
    email_column_index: int,
    remove_duplicates: bool,
    max_emails: Optional[int],
) -> List[str]:
    emails: List[str] = []
    seen = set()
    for row in rows:
        if email_column_index >= len(row):
            continue
        value = _normalize_email(row[email_column_index])
        if not value:
            continue
        key = value.lower()
        if remove_duplicates:
            if key in seen:
                continue
            seen.add(key)
        emails.append(value)
        if max_emails is not None and len(emails) > max_emails:
            raise FileProcessingError(
                "Email count exceeds the maximum allowed",
                details={"max_emails": max_emails, "seen": len(emails)},
            )
    return emails


def _detect_csv_dialect(sample: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(sample)
    except csv.Error as exc:
        logger.warning("file.csv.dialect_detect_failed", extra={"error": str(exc)})
        raise FileProcessingError("Unable to detect CSV delimiter. Please upload a standard comma-separated CSV.") from exc


def _parse_csv(
    data: bytes,
    email_column: str,
    first_row_has_labels: bool,
    remove_duplicates: bool,
    max_emails: Optional[int],
) -> ParsedEmails:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        logger.warning("file.csv.decode_failed", extra={"error": str(exc)})
        raise FileProcessingError("CSV must be UTF-8 encoded") from exc

    sample = text[:2048]
    dialect = _detect_csv_dialect(sample)
    reader = csv.reader(io.StringIO(text), dialect)
    first_row = next(reader, None)
    if first_row is None:
        raise FileProcessingError("CSV file is empty")
    headers = [str(value).strip() for value in first_row]
    email_column_index = _resolve_column_index(headers, email_column, first_row_has_labels)
    if email_column_index >= len(headers):
        raise FileProcessingError("Selected email column is outside the available columns")

    rows_iter: Iterable[Sequence[object]] = reader if first_row_has_labels else chain([first_row], reader)

    emails = _collect_emails(rows_iter, email_column_index, remove_duplicates, max_emails)
    if not emails:
        raise FileProcessingError("No emails found in the selected column")
    return ParsedEmails(emails=emails, email_column_index=email_column_index)


def _parse_xlsx(
    data: bytes,
    email_column: str,
    first_row_has_labels: bool,
    remove_duplicates: bool,
    max_emails: Optional[int],
) -> ParsedEmails:
    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if len(workbook.sheetnames) > 1:
        raise FileProcessingError("Multiple sheets detected. Please upload one sheet per file.")
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    first_row = next(rows, None)
    if first_row is None:
        raise FileProcessingError("Excel file is empty")
    headers = [str(value).strip() if value is not None else "" for value in first_row]
    email_column_index = _resolve_column_index(headers, email_column, first_row_has_labels)
    if email_column_index >= len(headers):
        raise FileProcessingError("Selected email column is outside the available columns")

    rows_iter = rows if first_row_has_labels else chain([first_row], rows)

    emails = _collect_emails(rows_iter, email_column_index, remove_duplicates, max_emails)
    if not emails:
        raise FileProcessingError("No emails found in the selected column")
    return ParsedEmails(emails=emails, email_column_index=email_column_index)


def _parse_xls(
    data: bytes,
    email_column: str,
    first_row_has_labels: bool,
    remove_duplicates: bool,
    max_emails: Optional[int],
) -> ParsedEmails:
    workbook = xlrd.open_workbook(file_contents=data)
    if workbook.nsheets > 1:
        raise FileProcessingError("Multiple sheets detected. Please upload one sheet per file.")
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise FileProcessingError("Excel file is empty")
    first_row = sheet.row_values(0)
    headers = [str(value).strip() if value is not None else "" for value in first_row]
    email_column_index = _resolve_column_index(headers, email_column, first_row_has_labels)
    if email_column_index >= len(headers):
        raise FileProcessingError("Selected email column is outside the available columns")

    rows: List[Sequence[object]] = []
    start_row = 1 if first_row_has_labels else 0
    for row_index in range(start_row, sheet.nrows):
        rows.append(sheet.row_values(row_index))

    emails = _collect_emails(rows, email_column_index, remove_duplicates, max_emails)
    if not emails:
        raise FileProcessingError("No emails found in the selected column")
    return ParsedEmails(emails=emails, email_column_index=email_column_index)


def parse_emails_from_upload(
    filename: str,
    data: bytes,
    email_column: str,
    first_row_has_labels: bool,
    remove_duplicates: bool,
    max_emails: Optional[int],
) -> ParsedEmails:
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise FileProcessingError("Unsupported file type", details={"extension": extension})
    if extension == ".csv":
        return _parse_csv(data, email_column, first_row_has_labels, remove_duplicates, max_emails)
    if extension == ".xlsx":
        return _parse_xlsx(data, email_column, first_row_has_labels, remove_duplicates, max_emails)
    return _parse_xls(data, email_column, first_row_has_labels, remove_duplicates, max_emails)


def _build_results_map(details: dict) -> dict:
    results: dict[str, dict] = {}
    for job in details.get("jobs") or []:
        email = job.get("email") or {}
        email_address = job.get("email_address") or email.get("email_address")
        if not email_address:
            continue
        key = str(email_address).strip().lower()
        if not key:
            continue
        results[key] = {
            "verification_status": email.get("status") or job.get("status"),
            "is_role_based": email.get("is_role_based"),
            "validated_at": email.get("validated_at"),
        }
    return results


def _append_output_columns(values: list, result: dict) -> list:
    return values + [
        result.get("verification_status") or "",
        "true" if result.get("is_role_based") is True else "false" if result.get("is_role_based") is False else "",
        result.get("validated_at") or "",
    ]


def write_verified_output(
    source_path: Path,
    output_path: Path,
    email_column_index: int,
    first_row_has_labels: bool,
    task_detail: dict,
) -> None:
    extension = source_path.suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise FileProcessingError("Unsupported file type", details={"extension": extension})

    results_map = _build_results_map(task_detail)
    if extension == ".csv":
        _write_csv_output(source_path, output_path, email_column_index, first_row_has_labels, results_map)
        return
    if extension == ".xlsx":
        _write_xlsx_output(source_path, output_path, email_column_index, first_row_has_labels, results_map)
        return
    _write_xls_output(source_path, output_path, email_column_index, first_row_has_labels, results_map)


def _write_csv_output(
    source_path: Path,
    output_path: Path,
    email_column_index: int,
    first_row_has_labels: bool,
    results_map: dict,
) -> None:
    text = source_path.read_text(encoding="utf-8-sig")
    dialect = _detect_csv_dialect(text[:2048])
    reader = csv.reader(io.StringIO(text), dialect)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as out_file:
        writer = csv.writer(out_file, dialect)
        for row_index, row in enumerate(reader):
            if row_index == 0 and first_row_has_labels:
                writer.writerow(row + OUTPUT_COLUMNS)
                continue
            email_value = row[email_column_index] if email_column_index < len(row) else ""
            key = str(email_value).strip().lower()
            result = results_map.get(key, {})
            writer.writerow(_append_output_columns(list(row), result))


def _write_xlsx_output(
    source_path: Path,
    output_path: Path,
    email_column_index: int,
    first_row_has_labels: bool,
    results_map: dict,
) -> None:
    workbook = openpyxl.load_workbook(source_path)
    if len(workbook.sheetnames) > 1:
        raise FileProcessingError("Multiple sheets detected. Please upload one sheet per file.")
    sheet = workbook[workbook.sheetnames[0]]
    start_col = sheet.max_column + 1
    if first_row_has_labels:
        for offset, column_name in enumerate(OUTPUT_COLUMNS):
            sheet.cell(row=1, column=start_col + offset, value=column_name)
    for row_idx in range(1 if not first_row_has_labels else 2, sheet.max_row + 1):
        email_cell = sheet.cell(row=row_idx, column=email_column_index + 1).value
        key = str(email_cell).strip().lower() if email_cell is not None else ""
        result = results_map.get(key, {})
        values = _append_output_columns([], result)
        for offset, value in enumerate(values):
            sheet.cell(row=row_idx, column=start_col + offset, value=value)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _write_xls_output(
    source_path: Path,
    output_path: Path,
    email_column_index: int,
    first_row_has_labels: bool,
    results_map: dict,
) -> None:
    workbook = xlrd.open_workbook(source_path)
    if workbook.nsheets > 1:
        raise FileProcessingError("Multiple sheets detected. Please upload one sheet per file.")
    sheet = workbook.sheet_by_index(0)
    out_book = xlwt.Workbook()
    out_sheet = out_book.add_sheet(sheet.name)

    start_col = sheet.ncols
    for row_idx in range(sheet.nrows):
        row_values = sheet.row_values(row_idx)
        for col_idx, value in enumerate(row_values):
            out_sheet.write(row_idx, col_idx, value)
        if row_idx == 0 and first_row_has_labels:
            for offset, column_name in enumerate(OUTPUT_COLUMNS):
                out_sheet.write(row_idx, start_col + offset, column_name)
            continue
        email_value = row_values[email_column_index] if email_column_index < len(row_values) else ""
        key = str(email_value).strip().lower() if email_value is not None else ""
        result = results_map.get(key, {})
        values = _append_output_columns([], result)
        for offset, value in enumerate(values):
            out_sheet.write(row_idx, start_col + offset, value)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_book.save(str(output_path))
